# coding:utf-8
# @Time : 2022/3/19 22:51
# @Author : cewinhot
# @Version：0.1
# @File : f4ratio2xlsx.py

import openpyxl
import sys
from openpyxl.styles import Font, Alignment

def detect_len(data, length):
    for i in range(LEN):
        len_tmp = len(data[i])
        if len_tmp > length[i]:
            length[i] = len_tmp

# Default setting
if __name__ == '__main__':
    path = sys.argv[1]
    head = ["A", "B", "X", "C", "O", "α", "std.err", "Z", "support"]
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(path.replace(".result",""))
    LEN = 9
    # 设置表头字体,字号,加粗,居中
    sheet.append(head)
    for i in range(LEN):
        sheet.cell(row=1, column=i + 1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=1, column=i + 1).font = Font(name='Times New Roman', size=14, bold=True)
    count = 2
    with open(path) as text:
        length = [0 for _ in range(LEN)]  # 保存列宽的列表
        for line in text:
            data = line.strip().split()
            detect_len(data, length)
            sheet.append(data)
            for i in range(LEN):
                sheet.cell(row=count, column=i + 1).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=count, column=i + 1).font = Font(name='Times New Roman', size=12)
            count += 1
        # 设置列宽
        for i in range(LEN):
            sheet.column_dimensions[chr(i + 65)].width = length[i] + 6
    del wb['Sheet']  # 删除原始表格
    wb.save(path.replace('result', 'xlsx'))
