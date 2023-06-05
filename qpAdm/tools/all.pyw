import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DIR = os.getcwd()
wbt = openpyxl.Workbook()
for root, dirs, files in os.walk(DIR):
    for file in files:
        if file.endswith('_result.xlsx'):
            t_name = file.split('_')[0]
            wst = wbt.create_sheet(t_name)

            wb = openpyxl.load_workbook(os.path.join(root, file))
            ws1 = wb['qpAdm']
            ws2 = wb['success']
            for row in ws1.rows:
                wst.append([cell.value for cell in row])

            wst.append(['', ''])

            for row in ws2.rows:
                wst.append([cell.value for cell in row])
            wb.close()

wbt.remove(wbt['Sheet'])
wbt.save('all.xlsx')
