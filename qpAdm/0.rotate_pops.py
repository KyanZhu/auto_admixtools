# coding:utf-8
# @Time : 2022/4/3 11:15
# @Author : cewinhot
# @Version：V 0.4
# @File : 0.rotate_pops.py


import os
import sys
from shutil import copy2
import openpyxl
from openpyxl.styles import Font, Alignment
from utils import *


def create_file(dir, page, tar_list, src_list, out_list):
    tools = ['0.qpAdm_bsub.template', '1.rotate_prepare.py', '2.bsub_qpAdm.sh',
             '3.grep_result.sh', '5.result2excel.py']
    page = 'src' + str(page)
    di = os.path.join(dir, page)
    if not os.path.exists(di):
        os.mkdir(di)
    # 复制工具
    for tool in tools:
        try:
            copy2(os.path.join(tool_path, tool), di)
        except:
            pass
    tar = open(os.path.join(di, '0.targets'), 'w')
    src = open(os.path.join(di, '0.sources'), 'w')
    out = open(os.path.join(di, '0.outgroups'), 'w')
    tar.write('\n'.join(convert(tar_list)))
    src.write('\n'.join(convert(src_list)))
    out.write('\n'.join(convert(out_list)))
    tar.close()
    src.close()
    out.close()


def create_qpadm(path, dir):
    wb = openpyxl.load_workbook(path)
    ws = wb['qpAdm']
    row = ws.max_row
    col = ws.max_column
    if len(wb.sheetnames) > 1:
        for i in wb.sheetnames[1:]:
            del wb[i]
    wc = wb.create_sheet('Content')
    wc.append(['Method', 'Sheet', 'Source1', 'Source2', 'Source3'])
    tar_list, src_list, out_list = [], [], []
    for i in range(1, row):
        b = [tar_list, src_list, out_list]
        for j, k in enumerate(b):
            value = str(ws.cell(row=i + 1, column=j + 1).value)
            if value != "None" and value not in k:
                k.append(value)
    # 检测是否有人群重合
    duplication_flag = False
    tar_set = set(remove_comment(convert(tar_list)))
    src_set = set(remove_comment(convert(src_list)))
    out_set = set(remove_comment(convert(out_list)))
    if (src_set & out_set):
        duplication = ", ".join(list(src_set & out_set))
        print(f"duplication groups [ {duplication} ] within: source and outgroup")
        duplication_flag = True
    if (src_set & tar_set):
        duplication = ", ".join(list(src_set & tar_set))
        print(f"duplication groups [ {duplication} ] within: source and target")
        duplication_flag = True
    if (tar_set & out_set):
        duplication = ", ".join(list(tar_set & out_set))
        print(f"duplication groups [ {duplication} ] within: target and outgroup")
        duplication_flag = True
    if duplication_flag:
        sys.exit()
    page = 1
    # qpWave检验source的异质性 pairwise/c2
    # combs = pairwise(src_list)  # 所有组合
    combs = c2_src(src_list)  # 不同组
    for comb in combs:
        ws = wb.create_sheet(str(page))
        head = ['Source1', 'Source2', 'Outgroups']
        ws.append(head)
        ws.cell(row=2, column=1).value = str(comb[0])
        ws.cell(row=2, column=2).value = str(comb[1])
        for i, k in enumerate(out_list):
            ws.cell(row=i + 2, column=3).value = str(k)
        create_file(dir, page, str(comb[0]), str(
            comb[1]), remove_comment(convert(out_list)))
        wc.append(['pairwise qpwave', page, comb[0], comb[1]])
        page += 1
    # generate check.poplist
    poplist = ' '.join(remove_comment(tar_list + src_list + out_list))
    with open(os.path.join(dir, 'check.poplist'), 'wt', encoding='utf-8') as f:
        f.write(poplist)
    # qpAdm rotation/combination
    for i, func in enumerate([c1, c2, c3]):
        rotate = True if '# Rotation' in src_list or '#Rotation' in src_list or '# rotation' in src_list or '#rotation' in src_list else False
        combs = func(src_list)  # get source combination list
        for comb in combs:
            sub_src_list = comb
            # if not is_include(comb, src_list, '# YR'): continue  # 检查特定人群是否存在, 存在继续执行, include(comb, src_li, target):
            # if not is_include(comb, src_list, '# Deep ancestry'): continue  # 检查特定人群是否存在, 存在继续执行, include(comb, src_li, target):
            if rotate:
                sub_out_list = out_list + list(set(src_list) - set(comb))
            else:
                sub_out_list = out_list
            ws = wb.create_sheet(str(page))
            head = ['Targets', 'Sources', 'Outgroups']
            ws.append(head)
            tar_list = remove_comment(convert(tar_list))
            sub_src_list = remove_comment(convert(sub_src_list))
            sub_out_list = remove_comment(convert(sub_out_list))
            for k, l in enumerate(tar_list):
                ws.cell(row=k + 2, column=1).value = l
            for k, l in enumerate(sub_src_list):
                ws.cell(row=k + 2, column=2).value = l
            for k, l in enumerate(sub_out_list):
                ws.cell(row=k + 2, column=3).value = l
            create_file(dir, page, tar_list, sub_src_list, sub_out_list)
            wc.append(['{}-way admixture'.format(i + 1), page] + sub_src_list)
            page += 1
    # 调整格式
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        row = ws.max_row
        col = ws.max_column
        # 居中对齐
        for j in range(col):
            ws.cell(row=1, column=j +
                    1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=j + 1).font = Font(bold=True,
                                                     size=13, name='Times New Roman')
            # bug: source人数多会报错, 未修复
            # for i in range(1, row):
            #     ws.cell(row=i + 1, column=j + 1).alignment = Alignment(horizontal='center', vertical='center')
            #     ws.cell(row=i + 1, column=j + 1).font = Font(size=12, name='Times New Roman')
        for j in range(3):
            ws.column_dimensions[chr(i + 65)].width = 35
        ws = wb['Content']
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 35
    wb.save(path)


def create_xlsx(path):
    wb = openpyxl.Workbook()
    head = ['Targets', 'Sources', 'Fixed outgroups']
    ws = wb.create_sheet('qpAdm')
    ws.append(head)
    for j in range(3):
        ws.cell(row=1, column=j +
                1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=1, column=j + 1).font = Font(bold=True,
                                                 size=13, name='Times New Roman')
        ws.column_dimensions[chr(j + 65)].width = 35
    del wb['Sheet']
    wb.save(path)


if __name__ == '__main__':
    tool_path = os.path.dirname(os.path.abspath(
        os.path.join(os.getcwd(), sys.argv[0])))
    dir = os.path.dirname(tool_path)
    pop_path = os.path.join(dir, os.path.basename(dir) + '_pops.xlsx')
    if os.path.exists(pop_path):
        create_qpadm(pop_path, dir)
    else:
        create_xlsx(pop_path)
