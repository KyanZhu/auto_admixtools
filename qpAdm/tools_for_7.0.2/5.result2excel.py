# coding:utf-8
# @Time : 2022/4/4 12:42
# @Author : cewinhot
# @Version：V 1.1
# @File : result2excel.py


import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# 自定义参数
HEAD = ["target", "source1", "source2", "source3", "ratio1", "ratio2", "ratio3", "std.error1", "std.error2",
        "std.error3", "tail", "totmean1", "totmean2", "totmean3"]

os.chdir(os.path.dirname(__file__))
DIR = os.getcwd()
INPUT_FILE = os.path.join(DIR, '4.all_result.txt')
INPUT_EXCEL = os.path.join(DIR, os.path.basename(DIR) + '_pops.xlsx')
INPUT_EXCEL = os.path.join(DIR, 'pops.xlsx')
OUTPUT_EXCEL = os.path.join(DIR, os.path.basename(DIR) + '_result.xlsx')
THRESH_STD = 1  # 0.15
THRESH_TAIL = 0.05
WARNING_COL = "ff5252"


def process1(li):
    new_li = []
    new_li.extend(li[0:4] + li[6:9] + li[11:14] +
                  [li[15][:6], li[17][:6], li[18][:6], li[19][:6]])
    return new_li


def detect_len(line, length, len1):
    for i in range(len1):
        t = len(line[i])
        if t > length[i]:
            length[i] = t


def fill_color(value):
    if value >= 3:
        return col[0]
    elif value >= 2.6:
        return col[1]
    elif value <= -3:
        return col[3]
    elif value <= -2.6:
        return col[2]
    else:
        return ""


def cell_search_text(cell, text):
    if text in cell.value:
        return True
    else:
        return False


def cell_pattern(cell, target=None, color=None, bold=False, size=None, font=None, italic=None, horizontal=None,
                 vertical=None):
    '''不定义target则直接定义表格样式
       若定义target则判断cell中是否包含target有则自定义表格样式'''
    if target == None or cell_search_text(cell, target):
        if color:
            cell.fill = PatternFill('solid', fgColor=color)
        if bold:
            cell.font = Font(bold=bold)
        if size:
            cell.font = Font(size=size)
        if font:
            cell.font = Font(name=font)
        if italic:
            cell.font = Font(italic=italic)
        if horizontal:
            cell.alignment = Alignment(horizontal=horizontal)
        if vertical:
            cell.alignment = Alignment(vertical=vertical)
        return


# 初始化
file = open(INPUT_FILE, 'r')
col = ["82b1ff", "81d4fa", "ffcc80", "ff8a80"]
wb = openpyxl.load_workbook(INPUT_EXCEL)
sheet_all = wb.create_sheet("all")
sheet_suc = wb.create_sheet("success")
for i in wb.sheetnames:
    if i.isdigit():
        del wb[i]  # 删除原始表格
sheet_all.append(HEAD)
sheet_suc.append(HEAD)
len1 = len(HEAD)
length_all = [len(x) for x in HEAD]
length_suc = [len(x) for x in HEAD]
for i in range(len1):
    sheet_all.cell(row=1, column=i +
                   1).alignment = Alignment(horizontal='center', vertical='center')
    sheet_all.cell(row=1, column=i +
                   1).font = Font(name='Times New Roman', size=14, bold=True)
    sheet_suc.cell(row=1, column=i +
                   1).alignment = Alignment(horizontal='center', vertical='center')
    sheet_suc.cell(row=1, column=i +
                   1).font = Font(name='Times New Roman', size=14, bold=True)
sheet_row = 2
sheet_suc_row = 2
for line in file:
    # 13 17 21 : 1,2,3-way
    line = line.split(' ')
    line = process1(line)
    sheet_all.append(line)
    detect_len(line, length_all, len1)  # 计算列宽
    for j in range(len1):  # 居中
        sheet_all.cell(row=sheet_row, column=j +
                       1).alignment = Alignment(horizontal='center', vertical='center')
    flag = 1  # 1 表示模拟成功
    # 检查coefficient
    c1, c2, c3 = sheet_all.cell(row=sheet_row, column=5), sheet_all.cell(row=sheet_row, column=6), sheet_all.cell(
        row=sheet_row,
        column=7)
    if '-' in c1.value:
        cell_pattern(c1, color=WARNING_COL)
        flag = 0
    if '-' in c2.value:
        cell_pattern(c2, color=WARNING_COL)
        flag = 0
    if '-' in c3.value:
        cell_pattern(c3, color=WARNING_COL)
        flag = 0
    # 检查std. errors
    e1, e2, e3 = sheet_all.cell(row=sheet_row, column=8), sheet_all.cell(row=sheet_row, column=9), sheet_all.cell(row=sheet_row, column=10)
    if e1.value != "":
        if float(e1.value) > THRESH_STD or float(e1.value) == 0:
            cell_pattern(e1, color=WARNING_COL)
            flag = 0
    if e2.value != "":
        if float(e2.value) > THRESH_STD:
            cell_pattern(e2, color=WARNING_COL)
            flag = 0
    if e3.value != "":
        if float(e3.value) > THRESH_STD:
            cell_pattern(e3, color=WARNING_COL)
            flag = 0
    # 检查std. 是否大于ratio
    if e1.value != "":
        if float(e1.value) > float(c1.value):
            cell_pattern(e1, color=WARNING_COL)
            flag = 0
    if e2.value != "":
        if float(e2.value) > float(c2.value):
            cell_pattern(e2, color=WARNING_COL)
            flag = 0
    if e3.value != "":
        if float(e3.value) > float(c3.value):
            cell_pattern(e3, color=WARNING_COL)
            flag = 0
    # 检查tail
    cell = sheet_all.cell(row=sheet_row, column=11)
    if cell.value != "":
        if float(cell.value) < THRESH_TAIL:
            cell_pattern(cell, color=WARNING_COL)
            flag = 0
    # 1-way情况下只要tail>0.05就成立
    if c2.value == "" and float(cell.value) > THRESH_TAIL:
        flag = 1
    if flag:  # 通过检查, 写入success
        detect_len(line, length_suc, len1)  # 计算列宽
        sheet_suc.append(line)
        for j in range(len1):  # 居中
            sheet_suc.cell(row=sheet_suc_row, column=j + 1).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
        sheet_suc_row += 1
    sheet_row += 1
for i in range(len1):
    sheet_all.column_dimensions[chr(i + 65)].width = length_all[i] + 3
    sheet_suc.column_dimensions[chr(i + 65)].width = length_suc[i] + 3
wb.save(OUTPUT_EXCEL)
fig = open(os.path.join(DIR, '6.plot.txt'), 'w')
fig.write("target.tail source Proportion\n")
for cell in sheet_suc:
    if cell[0].value != 'target':
        fig.write(cell[10].value + '|' + cell[0].value + ' ' +
                  cell[1].value + ' ' + cell[4].value + '\n')
        if cell[2].value != '':
            fig.write(cell[10].value + '|' + cell[0].value +
                      ' ' + cell[2].value + ' ' + cell[5].value + '\n')
        if cell[3].value != '':
            fig.write(cell[10].value + '|' + cell[0].value +
                      ' ' + cell[3].value + ' ' + cell[6].value + '\n')
fig.close()
