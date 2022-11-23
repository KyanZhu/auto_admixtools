# coding:utf-8
# @Time : 2021/9/25 12:42
# @Author : cewinhot
# @Version：V 1
# @File : beautyexcel.py


import os
import openpyxl
import argparse
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def is_digits(a):
    for i in a:
        if i not in "1234567890-+.":
            return False
    return True


def detect_len(line, length, len1):
    for i in range(len1):
        t = len(line[i])
        if t > length[i]:
            length[i] = t


def cell_search_text(cell, text):
    if type(cell.value) != str or text not in cell.value:
        return False
    else:
        return True


def cell_pattern(cell, target=None, color=None, bold=False, size=None, font=None, italic=None, horizontal=None,
                 vertical=None, bigger=None, smaller=None):
    '''不定义target则直接定义表格样式
       若定义target则判断cell中是否包含target有则自定义表格样式'''
    if bigger != None:
        if type(cell.value) == float or type(cell.value) == int or (type(cell.value) == str and is_digits(cell.value)):
            cell.fill = PatternFill('solid', fgColor=color)
            cell.font = Font(italic=italic, name=font, size=size, bold=bold)
            cell.alignment = Alignment(vertical=vertical, horizontal=horizontal)
        return
    if smaller != None:
        if type(cell.value) == float or type(cell.value) == int or (type(cell.value) == str and is_digits(cell.value)):
            if float(cell.value) <= smaller:
                cell.fill = PatternFill('solid', fgColor=color)
                cell.font = Font(italic=italic, name=font, size=size, bold=bold)
                cell.alignment = Alignment(vertical=vertical, horizontal=horizontal)
        return
    if target == None or cell_search_text(cell, target):
        if color:
            cell.fill = PatternFill('solid', fgColor=color)
        cell.font = Font(italic=italic, name=font, size=size, bold=bold)
        cell.alignment = Alignment(vertical=vertical, horizontal=horizontal)


def fill(workbook=None, sheet=None, target=None, row_start=None, row_end=None, col_start=None, col_end=None,
         color=None, bold=None, size=None, font=None, italic=None, horizontal=None, vertical=None, adjust=False,
         bigger=None, smaller=None):
    wb = workbook
    if sheet:
        sheets = [sheet] if type(sheet) == str else sheet
    else:
        sheets = wb.sheetnames
    for sheet in sheets:
        try:
            ws = wb[sheet]
        except:
            continue
        rs = ws.min_row if row_start == None else row_start
        re = ws.max_row if row_end == None else row_end
        cs = ws.min_column if col_start == None else col_start
        ce = ws.max_column if col_end == None else col_end
        cs = ord(col_start) - 64 if type(cs) == str else cs
        ce = ord(ce) - 64 if type(ce) == str else ce
        # cell_pattern
        for i in range(rs, re + 1):
            for j in range(cs, ce + 1):
                cell = ws.cell(row=i, column=j)
                cell_pattern(cell, target=target, color=color, bold=bold, size=size, font=font, italic=italic,
                             horizontal=horizontal, vertical=vertical, bigger=bigger, smaller=smaller)
        # 修改列宽
        if adjust:
            line = [0 for _ in range(ce - cs + 1)]
            print(ce, cs)
            print(line)
            for i in range(100):
                for j in range(ce - cs + 1):
                    try:
                        length = len(str(ws.cell(row=i + 1, column=cs + j).value))
                        print(length)
                        if length > line[j]:
                            line[j] = length
                    except:
                        pass
            print(line)
            for j in range(ce - cs + 1):
                if line[j] != 0:
                    ws.column_dimensions[chr(cs + j + 64)].width = line[j] + 3


if __name__ == '__main__':
    path = 'ancient.xlsx'
    wb = openpyxl.load_workbook(path)
    # sheet 对指定sheet进行填充(sheet名字符串或列表), None对所有sheet填充
    # target 对包含target字符进行填充
    # start,end指定其实行列, 列可用大写字母
    # color 指定填充颜色
    # adjust 自动调整列宽
    # 后面规则会覆盖前面规则!

    # 根据文字内容调整列宽, 对所有表, 第一行加粗, 居中
    fill(workbook=wb, sheet=None, target=None, row_start=1, row_end=1, col_start=None, col_end=None,
         color=None, bold=True, size=None, font='Times New Roman', italic=None, horizontal='center', vertical='center',
         adjust=True, bigger=None, smaller=None)

    # 从第二行开始, 对第6列小于<0的染色, 斜体
    fill(workbook=wb, sheet=None, target=None, row_start=2, row_end=None, col_start=6, col_end=6,
         color='ffc107', bold=None, size=None, font='Times New Roman', italic=True, horizontal='center',
         vertical='center', adjust=False, bigger=None, smaller=0)

    # 对1-400行, 第6列小于-2的染色, 加粗
    fill(workbook=wb, sheet=None, target=None, row_start=1, row_end=400, col_start='F', col_end='F',
         color='f44336', bold=True, size=None, font='Times New Roman', italic=None, horizontal='center',
         vertical='center', adjust=False, bigger=None, smaller=-2)

    # 对表Sheet1包含Malaysia的单元染色
    fill(workbook=wb, sheet='Sheet1', target='Malaysia', row_start=None, row_end=None, col_start=None, col_end=None,
         color='2d864b', bold=None, size=None, font='Times New Roman', italic=None, horizontal='center',
         vertical='center', adjust=False, bigger=None, smaller=None)
    wb.save(path.split('.xlsx')[0] + '_adjust.xlsx')
